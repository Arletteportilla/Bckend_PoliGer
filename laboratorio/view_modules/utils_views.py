"""
Vistas de utilidades y funciones auxiliares
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Avg, Count, Q
from datetime import datetime, timedelta
import json
import logging

from ..models import Germinacion, Polinizacion, Notification
from ..serializers import GerminacionSerializer, PolinizacionSerializer
# from ..permissions import require_reports_access, require_germinacion_access, require_polinizacion_access

logger = logging.getLogger(__name__)


# Funciones para generar reportes Excel
def apply_header_style(worksheet, row, columns):
    """Aplicar estilo de encabezado a una fila"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, columns + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    except ImportError:
        logger.warning("openpyxl no disponible para estilos de Excel")


def apply_data_style(worksheet, row, columns):
    """Aplicar estilo de datos a una fila"""
    try:
        from openpyxl.styles import Border, Side
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, columns + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
    except ImportError:
        logger.warning("openpyxl no disponible para estilos de Excel")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_germinaciones(request):
    """Generar reporte Excel de germinaciones"""
    try:
        # Usar ReportGenerator si está disponible
        try:
            from ..reports import ReportGenerator
            generator = ReportGenerator()
            return generator.generate_excel_report('germinaciones', {})
        except ImportError:
            return generar_reporte_basico_germinaciones(request)
    except Exception as e:
        logger.error(f"Error generando reporte de germinaciones: {e}")
        return JsonResponse({"error": f"Error generando reporte: {str(e)}"}, status=500)


def generar_reporte_basico_germinaciones(request):
    """Generar reporte Excel básico de germinaciones"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Germinaciones"
        
        # Encabezados
        headers = [
            'ID', 'Código', 'Género', 'Especie/Variedad', 'Fecha Siembra',
            'Cantidad Solicitada', 'No. Cápsulas', 'Estado Cápsulas',
            'Responsable', 'Fecha Creación'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        apply_header_style(ws, 1, len(headers))
        
        # Datos
        germinaciones = Germinacion.objects.all().order_by('-fecha_creacion')
        
        for row, germ in enumerate(germinaciones, 2):
            ws.cell(row=row, column=1, value=germ.id)
            ws.cell(row=row, column=2, value=germ.codigo or '')
            ws.cell(row=row, column=3, value=germ.genero or '')
            ws.cell(row=row, column=4, value=germ.especie_variedad or '')
            ws.cell(row=row, column=5, value=germ.fecha_siembra.strftime('%Y-%m-%d') if germ.fecha_siembra else '')
            ws.cell(row=row, column=6, value=germ.cantidad_solicitada or 0)
            ws.cell(row=row, column=7, value=germ.no_capsulas or 0)
            ws.cell(row=row, column=8, value=germ.estado_capsula or '')
            ws.cell(row=row, column=9, value=germ.responsable or '')
            ws.cell(row=row, column=10, value=germ.fecha_creacion.strftime('%Y-%m-%d %H:%M') if germ.fecha_creacion else '')
            
            apply_data_style(ws, row, len(headers))
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="germinaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error generando reporte básico de germinaciones: {e}")
        return JsonResponse({"error": f"Error generando reporte básico: {str(e)}"}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_polinizaciones(request):
    """Generar reporte Excel de polinizaciones"""
    try:
        # Usar ReportGenerator si está disponible
        try:
            from ..reports import ReportGenerator
            generator = ReportGenerator()
            return generator.generate_excel_report('polinizaciones', {})
        except ImportError:
            return generar_reporte_basico_polinizaciones(request)
    except Exception as e:
        logger.error(f"Error generando reporte de polinizaciones: {e}")
        return JsonResponse({"error": f"Error generando reporte: {str(e)}"}, status=500)


def generar_reporte_basico_polinizaciones(request):
    """Generar reporte Excel básico de polinizaciones"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Polinizaciones"
        
        # Encabezados
        headers = [
            'ID', 'Código', 'Género', 'Especie', 'Fecha Polinización',
            'Madre Género', 'Madre Especie', 'Padre Género', 'Padre Especie',
            'Nueva Género', 'Nueva Especie', 'Estado', 'Responsable', 'Fecha Creación'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        apply_header_style(ws, 1, len(headers))
        
        # Datos
        polinizaciones = Polinizacion.objects.all().order_by('-fecha_creacion')
        
        for row, pol in enumerate(polinizaciones, 2):
            ws.cell(row=row, column=1, value=pol.numero)
            ws.cell(row=row, column=2, value=pol.codigo or '')
            ws.cell(row=row, column=3, value=pol.genero or '')
            ws.cell(row=row, column=4, value=pol.especie or '')
            ws.cell(row=row, column=5, value=pol.fechapol.strftime('%Y-%m-%d') if pol.fechapol else '')
            ws.cell(row=row, column=6, value=pol.madre_genero or '')
            ws.cell(row=row, column=7, value=pol.madre_especie or '')
            ws.cell(row=row, column=8, value=pol.padre_genero or '')
            ws.cell(row=row, column=9, value=pol.padre_especie or '')
            ws.cell(row=row, column=10, value=pol.nueva_genero or '')
            ws.cell(row=row, column=11, value=pol.nueva_especie or '')
            ws.cell(row=row, column=12, value=pol.estado or '')
            ws.cell(row=row, column=13, value=pol.responsable or '')
            ws.cell(row=row, column=14, value=pol.fecha_creacion.strftime('%Y-%m-%d %H:%M') if pol.fecha_creacion else '')
            
            apply_data_style(ws, row, len(headers))
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="polinizaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error generando reporte básico de polinizaciones: {e}")
        return JsonResponse({"error": f"Error generando reporte básico: {str(e)}"}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_germinaciones(request):
    """Estadísticas de germinaciones"""
    try:
        # Totales por estado
        estados_count = Germinacion.objects.values('estado_capsula').annotate(
            count=Count('id')
        ).order_by('estado_capsula')
        
        # Totales por mes (últimos 12 meses)
        from django.db.models.functions import TruncMonth
        fecha_limite = datetime.now() - timedelta(days=365)
        por_mes_raw = Germinacion.objects.filter(
            fecha_creacion__gte=fecha_limite
        ).annotate(
            mes=TruncMonth('fecha_creacion')
        ).values('mes').annotate(
            total=Count('id')
        ).order_by('mes')
        
        # Formatear datos por mes para el frontend
        por_mes = []
        for item in por_mes_raw:
            por_mes.append({
                'mes': item['mes'].strftime('%Y-%m-%d') if item['mes'] else '',
                'total': item['total']
            })
        
        # Promedio de cápsulas
        promedio_capsulas = Germinacion.objects.aggregate(
            promedio=Avg('no_capsulas')
        )['promedio'] or 0
        
        # Calcular tasa de éxito (germinaciones completadas vs total)
        total_germinaciones = Germinacion.objects.count()
        germinaciones_exitosas = Germinacion.objects.filter(
            estado_capsula__in=['ABIERTA', 'GERMINADA']
        ).count()
        tasa_exito = round((germinaciones_exitosas / total_germinaciones * 100), 2) if total_germinaciones > 0 else 0
        
        # Promedio de días para germinar (simulado)
        promedio_dias_germinar = 15  # Valor por defecto, se puede calcular si hay datos de fechas
        
        # Top especies
        top_especies = Germinacion.objects.values('especie_variedad').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        return Response({
            'total': total_germinaciones,
            'tasa_exito': tasa_exito,
            'promedio_dias_germinar': promedio_dias_germinar,
            'promedio_capsulas': round(promedio_capsulas, 2),
            'estados': list(estados_count),
            'por_mes': por_mes,
            'top_especies': list(top_especies)
        })
        
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas de germinaciones: {e}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_polinizaciones(request):
    """Estadísticas de polinizaciones"""
    try:
        # Obtener filtros de fecha si existen
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Totales por estado
        estados_count = Polinizacion.objects.values('estado').annotate(
            count=Count('numero')
        ).order_by('estado')
        
        # Totales por mes basado en fechapol (fecha de polinización) o fechamad (fecha de maduración) como respaldo
        from django.db.models.functions import TruncMonth, Coalesce
        from django.utils.dateparse import parse_date
        
        # Usar fechapol para la tendencia
        query = Polinizacion.objects.filter(fechapol__isnull=False)

        # Aplicar filtros de fecha si existen
        if fecha_inicio:
            try:
                fecha_inicio_parsed = parse_date(fecha_inicio)
                if fecha_inicio_parsed:
                    query = query.filter(fechapol__gte=fecha_inicio_parsed)
            except Exception as e:
                logger.warning(f"Error parseando fecha_inicio: {e}")

        if fecha_fin:
            try:
                fecha_fin_parsed = parse_date(fecha_fin)
                if fecha_fin_parsed:
                    query = query.filter(fechapol__lte=fecha_fin_parsed)
            except Exception as e:
                logger.warning(f"Error parseando fecha_fin: {e}")

        # Si no hay filtros de fecha, usar los últimos 12 meses
        if not fecha_inicio and not fecha_fin:
            fecha_limite = datetime.now().date() - timedelta(days=365)
            query = query.filter(fechapol__gte=fecha_limite)

        # Log para debugging - contar antes de agrupar
        total_con_fecha = query.count()
        logger.info(f"Estadísticas polinizaciones - Total registros con fechapol: {total_con_fecha}")
        logger.info(f"Filtros aplicados - fecha_inicio: {fecha_inicio}, fecha_fin: {fecha_fin}")

        # Agrupar por mes usando fechapol
        por_mes_raw = query.annotate(
            mes=TruncMonth('fechapol')
        ).values('mes').annotate(
            total=Count('numero')
        ).order_by('mes')
        
        # Log para debugging
        logger.info(f"Estadísticas polinizaciones - Meses encontrados: {len(por_mes_raw)}")
        if por_mes_raw:
            logger.info(f"Estadísticas polinizaciones - Primeros meses: {list(por_mes_raw[:3])}")
        
        # Formatear datos por mes para el frontend
        por_mes = []
        for item in por_mes_raw:
            por_mes.append({
                'mes': item['mes'].strftime('%Y-%m-%d') if item['mes'] else '',
                'total': item['total']
            })
        
        # Calcular tasa de éxito (polinizaciones exitosas vs total)
        total_polinizaciones = Polinizacion.objects.count()
        polinizaciones_exitosas = Polinizacion.objects.filter(
            estado__in=['COMPLETADA', 'EXITOSA', 'MADURO']
        ).count()
        tasa_exito = round((polinizaciones_exitosas / total_polinizaciones * 100), 2) if total_polinizaciones > 0 else 0
        
        # Promedio de semillas por fruto (simulado)
        promedio_semillas_fruto = 25  # Valor por defecto, se puede calcular si hay datos
        
        # Top géneros
        top_generos = Polinizacion.objects.values('genero').annotate(
            count=Count('numero')
        ).order_by('-count')[:10]
        
        # Top especies
        top_especies = Polinizacion.objects.values('especie').annotate(
            count=Count('numero')
        ).order_by('-count')[:10]
        
        return Response({
            'total': total_polinizaciones,
            'tasa_exito': tasa_exito,
            'promedio_semillas_fruto': promedio_semillas_fruto,
            'estados': list(estados_count),
            'por_mes': por_mes,
            'top_generos': list(top_generos),
            'top_especies': list(top_especies)
        })
        
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas de polinizaciones: {e}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_usuario(request):
    """Estadísticas específicas del usuario logueado - Solo registros creados por el usuario"""
    try:
        user = request.user

        # Estadísticas de germinaciones creadas por el usuario
        mis_germinaciones = Germinacion.objects.filter(creado_por=user)
        total_germinaciones = mis_germinaciones.count()

        # Germinaciones actuales (en proceso, no finalizadas)
        germinaciones_actuales = mis_germinaciones.filter(
            estado_capsula__in=['Verde', 'En Proceso', 'Sin Fecha']
        ).exclude(
            estado_capsula__in=['Germinado', 'Finalizado']
        ).count()

        # Estadísticas de polinizaciones creadas por el usuario
        mis_polinizaciones = Polinizacion.objects.filter(creado_por=user)
        total_polinizaciones = mis_polinizaciones.count()

        # Polinizaciones actuales (en proceso, no finalizadas)
        polinizaciones_actuales = mis_polinizaciones.filter(
            estado__in=['INGRESADO', 'EN_PROCESO', 'GERMINANDO']
        ).exclude(
            estado__in=['COMPLETADA', 'FINALIZADA']
        ).count()

        # Polinizaciones completadas (para cálculo de éxito promedio)
        polinizaciones_completadas = mis_polinizaciones.filter(
            estado__in=['COMPLETADA', 'FINALIZADA', 'MADURO', 'LISTO']
        ).count()

        # Notificaciones no leídas
        notificaciones_no_leidas = Notification.objects.filter(
            usuario=user,
            leida=False
        ).count()

        # Respuesta en el formato esperado por el frontend
        return Response({
            'total_polinizaciones': total_polinizaciones,
            'total_germinaciones': total_germinaciones,
            'polinizaciones_actuales': polinizaciones_actuales,
            'germinaciones_actuales': germinaciones_actuales,
            'polinizaciones_completadas': polinizaciones_completadas,
            'usuario': user.username,
            'notificaciones_no_leidas': notificaciones_no_leidas
        })

    except Exception as e:
        logger.error(f"Error obteniendo estadísticas del usuario: {e}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_con_estadisticas(request):
    """Genera un reporte completo con estadísticas"""
    try:
        formato = request.GET.get('formato', 'excel').lower()
        incluir_estadisticas = request.GET.get('estadisticas', 'false').lower() == 'true'
        tipo_entidad = request.GET.get('tipo', 'ambos').lower()
        
        filtros = {
            'fecha_inicio': request.GET.get('fecha_inicio'),
            'fecha_fin': request.GET.get('fecha_fin'),
        }

        try:
            from ..reports import ReportGenerator
            generator = ReportGenerator()
            
            if incluir_estadisticas:
                if formato == 'pdf':
                    return generator.generate_pdf_report_with_stats(tipo_entidad, filtros)
                else:
                    return generator.generate_excel_report_with_stats(tipo_entidad, filtros)
            else:
                if formato == 'pdf':
                    return generator.generate_pdf_report(tipo_entidad, filtros)
                else:
                    return generator.generate_excel_report(tipo_entidad, filtros)

        except ImportError:
            logger.warning("ReportGenerator no disponible, usando reporte básico")
            if formato == 'pdf':
                return JsonResponse({
                    'error': 'Generación de PDF no disponible'
                }, status=500)
            else:
                return generar_reporte_basico_germinaciones(request)
                
    except Exception as e:
        logger.error(f"Error generando reporte con estadísticas: {e}")
        return Response({'error': str(e)}, status=500)