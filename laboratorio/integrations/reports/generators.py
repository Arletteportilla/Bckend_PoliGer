"""
Generador de reportes para el sistema de laboratorio
"""
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from datetime import datetime, timedelta
import io
from laboratorio.models import Germinacion, Polinizacion
from django.db.models import Count
from django.db.models.functions import TruncMonth


class ReportGenerator:
    """Generador de reportes en Excel y PDF"""
    
    def __init__(self):
        self.font_header = Font(bold=True, color="FFFFFF")
        self.fill_header = PatternFill(start_color="4A6CF7", end_color="4A6CF7", fill_type="solid")
        self.alignment_center = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def safe_get_value(self, obj, field_name, default=''):
        """Obtener valor de forma segura"""
        try:
            value = getattr(obj, field_name, default)
            return value if value is not None else default
        except:
            return default
    
    def format_date(self, date_value):
        """Formatear fecha de forma segura"""
        if date_value is None:
            return ''
        try:
            if hasattr(date_value, 'replace'):
                date_value = date_value.replace(tzinfo=None)
            return date_value.strftime('%Y-%m-%d') if date_value else ''
        except:
            return ''

    def generate_excel_report(self, data_type, filters=None):
        """Genera reporte Excel básico"""
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            
            if data_type == 'polinizaciones':
                return self._generate_polinizaciones_excel(ws, wb, filters)
            elif data_type == 'germinaciones':
                return self._generate_germinaciones_excel(ws, wb, filters)
            else:
                raise ValueError(f"Tipo de datos no soportado: {data_type}")
                
        except Exception as e:
            print(f"Error generando reporte Excel: {e}")
            raise

    def generate_pdf_report(self, data_type, filters=None):
        """Genera reporte PDF básico"""
        try:
            if data_type == 'polinizaciones':
                return self._generate_polinizaciones_pdf(filters)
            elif data_type == 'germinaciones':
                return self._generate_germinaciones_pdf(filters)
            else:
                raise ValueError(f"Tipo de datos no soportado: {data_type}")
                
        except Exception as e:
            print(f"Error generando reporte PDF: {e}")
            raise

    def _get_filtered_polinizaciones(self, filters=None):
        """Obtiene polinizaciones filtradas"""
        queryset = Polinizacion.objects.all()
        
        if filters:
            # Filtro por usuario actual
            if filters.get('usuario_actual'):
                from django.contrib.auth.models import User
                from django.db.models import Q
                username = filters['usuario_actual']

                try:
                    user = User.objects.get(username=username)

                    # Filtrar por creado_por o por responsable como fallback
                    queryset = queryset.filter(
                        Q(creado_por=user) |
                        Q(responsable__iexact=username) |
                        Q(responsable__icontains=f"{user.first_name} {user.last_name}".strip())
                    )
                except User.DoesNotExist:
                    # Si no se encuentra el usuario, devolver queryset vacío
                    queryset = queryset.none()
            
            # Filtro de búsqueda
            if filters.get('search'):
                from django.db.models import Q
                search = filters['search']
                queryset = queryset.filter(
                    Q(codigo__icontains=search) |
                    Q(genero__icontains=search) |
                    Q(especie__icontains=search) |
                    Q(madre_genero__icontains=search) |
                    Q(madre_especie__icontains=search) |
                    Q(padre_genero__icontains=search) |
                    Q(padre_especie__icontains=search) |
                    Q(nueva_genero__icontains=search) |
                    Q(nueva_especie__icontains=search) |
                    Q(ubicacion_nombre__icontains=search) |
                    Q(observaciones__icontains=search)
                )
            
            # Otros filtros
            if filters.get('fecha_inicio'):
                queryset = queryset.filter(fechapol__gte=filters['fecha_inicio'])
            if filters.get('fecha_fin'):
                queryset = queryset.filter(fechapol__lte=filters['fecha_fin'])
            if filters.get('estado'):
                queryset = queryset.filter(estado=filters['estado'])
        
        return queryset.order_by('-fecha_creacion')

    def _get_filtered_germinaciones(self, filters=None):
        """Obtiene germinaciones filtradas"""
        queryset = Germinacion.objects.all()
        
        if filters:
            # Filtro por usuario actual
            if filters.get('usuario_actual'):
                from django.contrib.auth.models import User
                from django.db.models import Q
                username = filters['usuario_actual']

                try:
                    user = User.objects.get(username=username)

                    # Filtrar por creado_por o por responsable como fallback
                    queryset = queryset.filter(
                        Q(creado_por=user) |
                        Q(responsable__iexact=username) |
                        Q(responsable__icontains=f"{user.first_name} {user.last_name}".strip())
                    )
                except User.DoesNotExist:
                    # Si no se encuentra el usuario, devolver queryset vacío
                    queryset = queryset.none()
            
            # Filtro de búsqueda
            if filters.get('search'):
                from django.db.models import Q
                search = filters['search']
                queryset = queryset.filter(
                    Q(codigo__icontains=search) |
                    Q(genero__icontains=search) |
                    Q(especie_variedad__icontains=search) |
                    Q(percha__icontains=search) |
                    Q(nivel__icontains=search) |
                    Q(observaciones__icontains=search) |
                    Q(responsable__icontains=search)
                )
            
            # Otros filtros
            if filters.get('fecha_inicio'):
                queryset = queryset.filter(fecha_siembra__gte=filters['fecha_inicio'])
            if filters.get('fecha_fin'):
                queryset = queryset.filter(fecha_siembra__lte=filters['fecha_fin'])
            if filters.get('etapa'):
                queryset = queryset.filter(etapa_actual=filters['etapa'])
        
        return queryset.order_by('-fecha_creacion')

    def _generate_polinizaciones_excel(self, ws, wb, filters=None):
        """Genera Excel de polinizaciones"""
        ws.title = "Polinizaciones"
        
        # Obtener datos filtrados
        polinizaciones = self._get_filtered_polinizaciones(filters)
        
        # Encabezados
        headers = [
            'Código', 'Género', 'Especie', 'Tipo', 'Fecha Polinización',
            'Fecha Maduración', 'Estado', 'Cantidad', 'Ubicación', 'Responsable'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = self.alignment_center
            cell.border = self.border
        
        # Escribir datos
        for row, pol in enumerate(polinizaciones, 2):
            ws.cell(row=row, column=1, value=self.safe_get_value(pol, 'codigo'))
            ws.cell(row=row, column=2, value=self.safe_get_value(pol, 'genero'))
            ws.cell(row=row, column=3, value=self.safe_get_value(pol, 'especie'))
            ws.cell(row=row, column=4, value=self.safe_get_value(pol, 'tipo_polinizacion'))
            ws.cell(row=row, column=5, value=self.format_date(pol.fechapol))
            ws.cell(row=row, column=6, value=self.format_date(pol.fechamad))
            ws.cell(row=row, column=7, value=self.safe_get_value(pol, 'estado'))
            ws.cell(row=row, column=8, value=self.safe_get_value(pol, 'cantidad_capsulas'))
            ws.cell(row=row, column=9, value=self.safe_get_value(pol, 'ubicacion_nombre'))
            ws.cell(row=row, column=10, value=self.safe_get_value(pol, 'responsable'))
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="polinizaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Guardar workbook en respuesta
        wb.save(response)
        return response

    def _generate_germinaciones_excel(self, ws, wb, filters=None):
        """Genera Excel de germinaciones"""
        ws.title = "Germinaciones"
        
        # Obtener datos filtrados
        germinaciones = self._get_filtered_germinaciones(filters)
        
        # Encabezados
        headers = [
            'Código', 'Género', 'Especie/Variedad', 'Fecha Siembra',
            'Ubicación', 'Cantidad Solicitada', 'No. Cápsulas',
            'Estado Cápsula', 'Estado Semilla', 'Responsable'
        ]

        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = self.alignment_center
            cell.border = self.border

        # Escribir datos
        for row, germ in enumerate(germinaciones, 2):
            # Construir ubicación concatenando percha y nivel
            percha = self.safe_get_value(germ, 'percha', '')
            nivel = self.safe_get_value(germ, 'nivel', '')
            ubicacion = f"{percha} {nivel}".strip() if percha or nivel else ''

            ws.cell(row=row, column=1, value=self.safe_get_value(germ, 'codigo'))
            ws.cell(row=row, column=2, value=self.safe_get_value(germ, 'genero'))
            ws.cell(row=row, column=3, value=self.safe_get_value(germ, 'especie_variedad'))
            ws.cell(row=row, column=4, value=self.format_date(germ.fecha_siembra))
            ws.cell(row=row, column=5, value=ubicacion)
            ws.cell(row=row, column=6, value=self.safe_get_value(germ, 'cantidad_solicitada'))
            ws.cell(row=row, column=7, value=self.safe_get_value(germ, 'no_capsulas'))
            ws.cell(row=row, column=8, value=self.safe_get_value(germ, 'estado_capsula'))
            ws.cell(row=row, column=9, value=self.safe_get_value(germ, 'estado_semilla'))
            ws.cell(row=row, column=10, value=self.safe_get_value(germ, 'responsable'))
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="germinaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Guardar workbook en respuesta
        wb.save(response)
        return response

    def _generate_polinizaciones_pdf(self, filters=None):
        """Genera PDF profesional de polinizaciones con tablas"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.enums import TA_CENTER
            import os
        except ImportError:
            raise ImportError("ReportLab no está instalado. Instala con: pip install reportlab")

        # Obtener datos filtrados
        polinizaciones = self._get_filtered_polinizaciones(filters)

        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="polinizaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        # Crear PDF con platypus en formato A4 horizontal
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                              rightMargin=30, leftMargin=30,
                              topMargin=50, bottomMargin=30)

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Contenido del PDF
        story = []

        # Logo Ecuagenera
        logo_path = r'C:\Users\arlet\Desktop\78\78\PoliGer\assets\images\Ecuagenera.png'
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=1.2*inch, height=1.2*inch)
                logo.hAlign = 'RIGHT'
                story.append(logo)
                story.append(Spacer(1, 0.1*inch))
            except Exception as e:
                print(f"No se pudo cargar el logo: {e}")

        # Título principal
        story.append(Paragraph("POLIGER ECUAGENERA", title_style))
        story.append(Paragraph("Reporte de Polinizaciones", subtitle_style))
        story.append(Spacer(1, 0.2*inch))

        # Información del reporte
        info_text = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if filters and filters.get('fecha_inicio'):
            info_text += f" | Desde: {filters['fecha_inicio']}"
        if filters and filters.get('fecha_fin'):
            info_text += f" hasta: {filters['fecha_fin']}"
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Preparar datos para la tabla principal
        data = [['Código', 'Tipo', 'Fecha\nPolini.', 'Fecha\nMad.', 'Nueva\nGénero',
                 'Nueva\nEspecie', 'Ubicación', 'Cantidad\nSolicitada', 'Cantidad\nDisponible']]

        # Contar por estado
        completadas = 0
        en_proceso = 0

        for pol in polinizaciones:
            estado = str(pol.estado or 'N/A')
            if estado in ['COMPLETADA', 'FINALIZADA', 'MADURO', 'LISTO', 'FINALIZADO']:
                completadas += 1
            else:
                en_proceso += 1

            # Ubicación combinada
            ubicacion = f"{pol.ubicacion_tipo or ''} {pol.ubicacion_nombre or ''}".strip() or 'N/A'

            data.append([
                str(pol.codigo or '')[:15],
                str(pol.tipo_polinizacion or 'N/A')[:10],
                str(pol.fechapol)[:10] if pol.fechapol else 'N/A',
                str(pol.fechamad)[:10] if pol.fechamad else 'N/A',
                str(pol.nueva_genero or pol.genero or '')[:12],
                str(pol.nueva_especie or pol.especie or '')[:15],
                ubicacion[:12],
                str(pol.cantidad_solicitada or '0'),
                str(pol.cantidad_disponible or '0')
            ])

        # Crear tabla principal (9 columnas)
        col_widths = [0.9*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.9*inch,
                     1.1*inch, 0.9*inch, 0.8*inch, 0.8*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Estilo de la tabla principal
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Tipo centrado
            ('ALIGN', (7, 1), (8, -1), 'CENTER'),  # Cantidades centradas
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),
        ]))

        story.append(table)
        story.append(Spacer(1, 0.4*inch))

        # Tabla de resumen
        total = completadas + en_proceso
        summary_data = [
            ['Estado', 'Cantidad'],
            ['Completadas', str(completadas)],
            ['En Proceso', str(en_proceso)],
            ['TOTAL', str(total)]
        ]

        summary_table = Table(summary_data, colWidths=[4*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),

            # Completadas (verde)
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#d1fae5')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),

            # En Proceso (amarillo)
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#fef3c7')),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.black),

            # Total (gris)
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 3), (-1, 3), colors.black),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),

            # General
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        story.append(summary_table)

        # Construir PDF
        doc.build(story)

        # Obtener PDF del buffer
        pdf_data = buffer.getvalue()
        buffer.close()

        response.write(pdf_data)
        return response

    def _generate_germinaciones_pdf(self, filters=None):
        """Genera PDF profesional de germinaciones con tablas"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            import os
        except ImportError:
            raise ImportError("ReportLab no está instalado. Instala con: pip install reportlab")

        # Obtener datos filtrados
        germinaciones = self._get_filtered_germinaciones(filters)

        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="germinaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        # Crear PDF con platypus en formato A4 horizontal
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                              rightMargin=30, leftMargin=30,
                              topMargin=50, bottomMargin=30)

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Contenido del PDF
        story = []

        # Logo Ecuagenera
        logo_path = r'C:\Users\arlet\Desktop\78\78\PoliGer\assets\images\Ecuagenera.png'
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=1.2*inch, height=1.2*inch)
                logo.hAlign = 'RIGHT'
                story.append(logo)
                story.append(Spacer(1, 0.1*inch))
            except Exception as e:
                print(f"No se pudo cargar el logo: {e}")

        # Título principal
        story.append(Paragraph("POLIGER ECUAGENERA", title_style))
        story.append(Paragraph("Reporte de Germinaciones", subtitle_style))
        story.append(Spacer(1, 0.2*inch))

        # Información del reporte
        info_text = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if filters and filters.get('fecha_inicio'):
            info_text += f" | Desde: {filters['fecha_inicio']}"
        if filters and filters.get('fecha_fin'):
            info_text += f" hasta: {filters['fecha_fin']}"
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Preparar datos para la tabla principal
        data = [['Código', 'Género', 'Especie/Variedad', 'Fecha\nSiembra', 'Cant.\nSolic.',
                 'Cápsulas', 'Estado', 'Clima', 'Responsable']]

        # Contar completadas y pendientes
        completadas = 0
        pendientes = 0

        for germ in germinaciones:
            # Determinar si está completada (tiene fecha de germinación)
            estado_germ = str(germ.estado_germinacion or germ.etapa_actual or 'N/A')
            if germ.fecha_germinacion or estado_germ in ['FINALIZADO', 'LISTA', 'FINALIZADA']:
                completadas += 1
            else:
                pendientes += 1

            data.append([
                str(germ.codigo or '')[:15],
                str(germ.genero or '')[:15],
                str(germ.especie_variedad or '')[:20],
                str(germ.fecha_siembra)[:10] if germ.fecha_siembra else 'N/A',
                str(germ.cantidad_solicitada or '0'),
                str(germ.no_capsulas or '0'),
                estado_germ[:12],
                str(germ.clima or 'I'),
                str(germ.responsable or '')[:15]
            ])

        # Crear tabla principal
        col_widths = [0.9*inch, 0.9*inch, 1.3*inch, 0.8*inch, 0.6*inch,
                     0.7*inch, 0.9*inch, 0.6*inch, 1.1*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Estilo de la tabla principal
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (5, -1), 'CENTER'),  # Cant. y Cápsulas centradas
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),
        ]))

        story.append(table)
        story.append(Spacer(1, 0.4*inch))

        # Tabla de resumen
        total = completadas + pendientes
        summary_data = [
            ['Estado', 'Cantidad'],
            ['Completadas (con fecha de germinación)', str(completadas)],
            ['Pendientes (sin fecha de germinación)', str(pendientes)],
            ['TOTAL', str(total)]
        ]

        summary_table = Table(summary_data, colWidths=[4*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),

            # Completadas (verde)
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#d1fae5')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),

            # Pendientes (amarillo)
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#fef3c7')),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.black),

            # Total (gris)
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 3), (-1, 3), colors.black),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),

            # General
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        story.append(summary_table)

        # Construir PDF
        doc.build(story)

        # Obtener PDF del buffer
        pdf_data = buffer.getvalue()
        buffer.close()

        response.write(pdf_data)
        return response
    
    def generate_excel_report_with_stats(self, tipo, filters=None):
        """Genera reporte Excel con estadísticas y gráficos"""
        try:
            wb = Workbook()
            
            if tipo == 'germinaciones':
                ws_data = wb.active
                ws_data.title = "Datos Germinaciones"
                self._fill_germinaciones_data(ws_data, filters)
                
                ws_stats = wb.create_sheet("Estadísticas")
                self._generate_estadisticas_germinaciones_sheet(ws_stats)
                
                ws_charts = wb.create_sheet("Gráficos")
                self._generate_charts_germinaciones_sheet(ws_charts)
                
            elif tipo == 'polinizaciones':
                ws_data = wb.active
                ws_data.title = "Datos Polinizaciones"
                self._fill_polinizaciones_data(ws_data, filters)
                
                ws_stats = wb.create_sheet("Estadísticas")
                self._generate_estadisticas_polinizaciones_sheet(ws_stats)
                
                ws_charts = wb.create_sheet("Gráficos")
                self._generate_charts_polinizaciones_sheet(ws_charts)
                
            elif tipo == 'ambos':
                ws_germ = wb.active
                ws_germ.title = "Datos Germinaciones"
                self._fill_germinaciones_data(ws_germ, filters)
                
                ws_pol = wb.create_sheet("Datos Polinizaciones")
                self._fill_polinizaciones_data(ws_pol, filters)
                
                ws_stats = wb.create_sheet("Estadísticas Generales")
                self._generate_estadisticas_generales_sheet(ws_stats)
                
                ws_charts = wb.create_sheet("Gráficos Generales")
                self._generate_charts_generales_sheet(ws_charts)
            
            return self._create_excel_response(wb, f'{tipo}_con_estadisticas')
            
        except Exception as e:
            raise Exception(f"Error generando reporte: {str(e)}")

    def _fill_germinaciones_data(self, ws, filters):
        """Llena una hoja con datos de germinaciones"""
        queryset = Germinacion.objects.select_related('creado_por', 'polinizacion').all()
        
        if filters:
            if filters.get('fecha_inicio'):
                queryset = queryset.filter(fecha_creacion__gte=filters['fecha_inicio'])
            if filters.get('fecha_fin'):
                queryset = queryset.filter(fecha_creacion__lte=filters['fecha_fin'])
        
        headers = [
            'ID', 'Código', 'Especie/Variedad', 'Fecha Polinización', 'Fecha Siembra',
            'Clima', 'Ubicación', 'Clima Lab', 'Cantidad Solicitada', 'No. Cápsulas',
            'Estado Cápsula', 'Estado Semilla', 'Cantidad Semilla', 'Semilla en Stock',
            'Observaciones', 'Responsable', 'Fecha Creación', 'Fecha Actualización',
            'Creado por', 'Fecha Germinación', 'Tipo Polinización', 'Etapa Actual'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row, germinacion in enumerate(queryset, 2):
            # Construir ubicación concatenando percha y nivel
            percha = self.safe_get_value(germinacion, 'percha', '')
            nivel = self.safe_get_value(germinacion, 'nivel', '')
            ubicacion = f"{percha} {nivel}".strip() if percha or nivel else ''

            ws.cell(row=row, column=1, value=self.safe_get_value(germinacion, 'id'))
            ws.cell(row=row, column=2, value=self.safe_get_value(germinacion, 'codigo'))
            ws.cell(row=row, column=3, value=self.safe_get_value(germinacion, 'especie_variedad'))
            ws.cell(row=row, column=4, value=self.format_date(self.safe_get_value(germinacion, 'fecha_polinizacion')))
            ws.cell(row=row, column=5, value=self.format_date(self.safe_get_value(germinacion, 'fecha_siembra')))
            ws.cell(row=row, column=6, value=self.safe_get_value(germinacion, 'clima'))
            ws.cell(row=row, column=7, value=ubicacion)
            ws.cell(row=row, column=8, value=self.safe_get_value(germinacion, 'clima_lab'))
            ws.cell(row=row, column=9, value=self.safe_get_value(germinacion, 'cantidad_solicitada'))
            ws.cell(row=row, column=10, value=self.safe_get_value(germinacion, 'no_capsulas'))
            ws.cell(row=row, column=11, value=self.safe_get_value(germinacion, 'estado_capsula'))
            ws.cell(row=row, column=12, value=self.safe_get_value(germinacion, 'estado_semilla'))
            ws.cell(row=row, column=13, value=self.safe_get_value(germinacion, 'cantidad_semilla'))
            ws.cell(row=row, column=14, value='Sí' if self.safe_get_value(germinacion, 'semilla_en_stock') else 'No')
            ws.cell(row=row, column=15, value=self.safe_get_value(germinacion, 'observaciones'))
            ws.cell(row=row, column=16, value=self.safe_get_value(germinacion, 'responsable'))
            ws.cell(row=row, column=17, value=self.format_date(self.safe_get_value(germinacion, 'fecha_creacion')))
            ws.cell(row=row, column=18, value=self.format_date(self.safe_get_value(germinacion, 'fecha_actualizacion')))
            ws.cell(row=row, column=19, value=self.safe_get_value(germinacion.creado_por, 'username') if germinacion.creado_por else '')
            ws.cell(row=row, column=20, value=self.format_date(self.safe_get_value(germinacion, 'fecha_germinacion')))
            ws.cell(row=row, column=21, value=self.safe_get_value(germinacion, 'tipo_polinizacion'))
            ws.cell(row=row, column=22, value=self.safe_get_value(germinacion, 'etapa_actual'))
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _fill_polinizaciones_data(self, ws, filters):
        """Llena una hoja con datos de polinizaciones"""
        queryset = Polinizacion.objects.select_related('creado_por').all()
        
        if filters:
            if filters.get('fecha_inicio'):
                queryset = queryset.filter(fecha_creacion__gte=filters['fecha_inicio'])
            if filters.get('fecha_fin'):
                queryset = queryset.filter(fecha_creacion__lte=filters['fecha_fin'])
        
        headers = [
            'Número', 'Código', 'Fecha Polinización', 'Fecha Maduración', 'Tipo Polinización',
            'Madre Código', 'Madre Clima', 'Padre Código', 'Padre Clima',
            'Ubicación Tipo', 'Ubicación Nombre', 'Cantidad Cápsulas', 'Responsable',
            'Disponible', 'Estado', 'Fecha Creación', 'Fecha Actualización', 'Creado por'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row, polinizacion in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=self.safe_get_value(polinizacion, 'numero'))
            ws.cell(row=row, column=2, value=self.safe_get_value(polinizacion, 'codigo'))
            ws.cell(row=row, column=3, value=self.format_date(self.safe_get_value(polinizacion, 'fechapol')))
            ws.cell(row=row, column=4, value=self.format_date(self.safe_get_value(polinizacion, 'fechamad')))
            ws.cell(row=row, column=5, value=self.safe_get_value(polinizacion, 'tipo_polinizacion'))
            ws.cell(row=row, column=6, value=self.safe_get_value(polinizacion, 'madre_codigo'))
            ws.cell(row=row, column=7, value=self.safe_get_value(polinizacion, 'madre_clima'))
            ws.cell(row=row, column=8, value=self.safe_get_value(polinizacion, 'padre_codigo'))
            ws.cell(row=row, column=9, value=self.safe_get_value(polinizacion, 'padre_clima'))
            ws.cell(row=row, column=10, value=self.safe_get_value(polinizacion, 'ubicacion_tipo'))
            ws.cell(row=row, column=11, value=self.safe_get_value(polinizacion, 'ubicacion_nombre'))
            ws.cell(row=row, column=12, value=self.safe_get_value(polinizacion, 'cantidad_capsulas'))
            ws.cell(row=row, column=13, value=self.safe_get_value(polinizacion, 'responsable'))
            ws.cell(row=row, column=14, value='Sí' if self.safe_get_value(polinizacion, 'disponible') else 'No')
            ws.cell(row=row, column=15, value=self.safe_get_value(polinizacion, 'estado'))
            ws.cell(row=row, column=16, value=self.format_date(self.safe_get_value(polinizacion, 'fecha_creacion')))
            ws.cell(row=row, column=17, value=self.format_date(self.safe_get_value(polinizacion, 'fecha_actualizacion')))
            ws.cell(row=row, column=18, value=self.safe_get_value(polinizacion.creado_por, 'username') if polinizacion.creado_por else '')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _generate_estadisticas_germinaciones_sheet(self, ws):
        """Genera hoja de estadísticas para germinaciones"""
        ws.title = "Estadísticas Germinaciones"
        
        ws.cell(row=1, column=1, value="ESTADÍSTICAS DE GERMINACIONES")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        germinaciones = Germinacion.objects.all()
        total_germinaciones = germinaciones.count()
        
        row = 3
        ws.cell(row=row, column=1, value="Total de Germinaciones:")
        ws.cell(row=row, column=2, value=total_germinaciones)
        
        row += 1
        ws.cell(row=row, column=1, value="Germinaciones con Semilla en Stock:")
        ws.cell(row=row, column=2, value=germinaciones.filter(semilla_en_stock=True).count())
        
        row += 1
        ws.cell(row=row, column=1, value="Germinaciones sin Semilla en Stock:")
        ws.cell(row=row, column=2, value=germinaciones.filter(semilla_en_stock=False).count())
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _generate_estadisticas_polinizaciones_sheet(self, ws):
        """Genera hoja de estadísticas para polinizaciones"""
        ws.title = "Estadísticas Polinizaciones"
        
        ws.cell(row=1, column=1, value="ESTADÍSTICAS DE POLINIZACIONES")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        polinizaciones = Polinizacion.objects.all()
        total_polinizaciones = polinizaciones.count()
        
        row = 3
        ws.cell(row=row, column=1, value="Total de Polinizaciones:")
        ws.cell(row=row, column=2, value=total_polinizaciones)
        
        row += 1
        ws.cell(row=row, column=1, value="Polinizaciones Disponibles:")
        ws.cell(row=row, column=2, value=polinizaciones.filter(disponible=True).count())
        
        row += 1
        ws.cell(row=row, column=1, value="Polinizaciones No Disponibles:")
        ws.cell(row=row, column=2, value=polinizaciones.filter(disponible=False).count())
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

    def _generate_estadisticas_generales_sheet(self, ws):
        """Genera hoja de estadísticas generales"""
        ws.title = "Estadísticas Generales"
        
        ws.cell(row=1, column=1, value="ESTADÍSTICAS GENERALES DEL SISTEMA")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        germinaciones = Germinacion.objects.all()
        polinizaciones = Polinizacion.objects.all()
        
        row = 3
        ws.cell(row=row, column=1, value="Total de Germinaciones:")
        ws.cell(row=row, column=2, value=germinaciones.count())
        
        row += 1
        ws.cell(row=row, column=1, value="Total de Polinizaciones:")
        ws.cell(row=row, column=2, value=polinizaciones.count())
        
        row += 1
        ws.cell(row=row, column=1, value="Total de Registros:")
        ws.cell(row=row, column=2, value=germinaciones.count() + polinizaciones.count())
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15

    def _generate_charts_germinaciones_sheet(self, ws):
        """Genera hoja de gráficos para germinaciones"""
        ws.title = "Gráficos Germinaciones"
        
        ws.cell(row=1, column=1, value="GRÁFICOS DE GERMINACIONES")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Gráfico simple de distribución por etapa
        etapas = Germinacion.objects.values('etapa_actual').annotate(count=Count('id'))
        
        ws.cell(row=3, column=1, value="Distribución por Etapa")
        ws.cell(row=3, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=5, column=1, value="Etapa")
        ws.cell(row=5, column=2, value="Cantidad")
        
        row = 6
        for etapa in etapas:
            if etapa['etapa_actual']:
                ws.cell(row=row, column=1, value=etapa['etapa_actual'])
                ws.cell(row=row, column=2, value=etapa['count'])
                row += 1
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.title = "Distribución por Etapa"
        chart.x_axis.title = "Etapa"
        chart.y_axis.title = "Cantidad"
        
        data = Reference(ws, min_col=2, min_row=5, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=6, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D10")

    def _generate_charts_polinizaciones_sheet(self, ws):
        """Genera hoja de gráficos para polinizaciones"""
        ws.title = "Gráficos Polinizaciones"
        
        ws.cell(row=1, column=1, value="GRÁFICOS DE POLINIZACIONES")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Gráfico simple de distribución por estado
        estados = Polinizacion.objects.values('estado').annotate(count=Count('id'))
        
        ws.cell(row=3, column=1, value="Distribución por Estado")
        ws.cell(row=3, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=5, column=1, value="Estado")
        ws.cell(row=5, column=2, value="Cantidad")
        
        row = 6
        for estado in estados:
            if estado['estado']:
                ws.cell(row=row, column=1, value=estado['estado'])
                ws.cell(row=row, column=2, value=estado['count'])
                row += 1
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.title = "Distribución por Estado"
        chart.x_axis.title = "Estado"
        chart.y_axis.title = "Cantidad"
        
        data = Reference(ws, min_col=2, min_row=5, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=6, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D10")

    def _generate_charts_generales_sheet(self, ws):
        """Genera hoja de gráficos generales"""
        ws.title = "Gráficos Generales"
        
        ws.cell(row=1, column=1, value="GRÁFICOS GENERALES DEL SISTEMA")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Gráfico de comparación general
        total_germinaciones = Germinacion.objects.count()
        total_polinizaciones = Polinizacion.objects.count()
        
        ws.cell(row=3, column=1, value="Comparación Germinaciones vs Polinizaciones")
        ws.cell(row=3, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=5, column=1, value="Tipo")
        ws.cell(row=5, column=2, value="Cantidad")
        
        ws.cell(row=6, column=1, value="Germinaciones")
        ws.cell(row=6, column=2, value=total_germinaciones)
        
        ws.cell(row=7, column=1, value="Polinizaciones")
        ws.cell(row=7, column=2, value=total_polinizaciones)
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.title = "Comparación General"
        chart.x_axis.title = "Tipo"
        chart.y_axis.title = "Cantidad"
        
        data = Reference(ws, min_col=2, min_row=5, max_row=7)
        cats = Reference(ws, min_col=1, min_row=6, max_row=7)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D10")

    def _create_excel_response(self, wb, report_type):
        """Crea respuesta HTTP para Excel"""
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'reporte_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    def generate_pdf_report_with_stats(self, tipo, filters=None):
        """Genera reporte PDF con estadísticas"""
        try:
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter

            # Encabezado poliger ecuagenera
            p.setFont("Helvetica-Bold", 18)
            p.drawCentredString(width / 2, height - 30, "POLIGER ECUAGENERA")

            p.setFont("Helvetica-Bold", 16)
            p.drawCentredString(width/2, height-60, f"REPORTE DE {tipo.upper()}")
            
            if tipo == 'germinaciones':
                queryset = Germinacion.objects.all()
                if filters:
                    if filters.get('fecha_inicio'):
                        queryset = queryset.filter(fecha_creacion__gte=filters['fecha_inicio'])
                    if filters.get('fecha_fin'):
                        queryset = queryset.filter(fecha_creacion__lte=filters['fecha_fin'])

                y_position = height - 110
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, y_position, "ESTADÍSTICAS DE GERMINACIONES")
                
                y_position -= 30
                p.setFont("Helvetica", 10)
                p.drawString(50, y_position, f"Total de Germinaciones: {queryset.count()}")
                
                y_position -= 20
                p.drawString(50, y_position, f"Con Semilla en Stock: {queryset.filter(semilla_en_stock=True).count()}")
                
                y_position -= 20
                p.drawString(50, y_position, f"Sin Semilla en Stock: {queryset.filter(semilla_en_stock=False).count()}")
                
            elif tipo == 'polinizaciones':
                queryset = Polinizacion.objects.all()
                if filters:
                    if filters.get('fecha_inicio'):
                        queryset = queryset.filter(fecha_creacion__gte=filters['fecha_inicio'])
                    if filters.get('fecha_fin'):
                        queryset = queryset.filter(fecha_creacion__lte=filters['fecha_fin'])

                y_position = height - 110
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, y_position, "ESTADÍSTICAS DE POLINIZACIONES")
                
                y_position -= 30
                p.setFont("Helvetica", 10)
                p.drawString(50, y_position, f"Total de Polinizaciones: {queryset.count()}")
                
                y_position -= 20
                p.drawString(50, y_position, f"Disponibles: {queryset.filter(disponible=True).count()}")
                
                y_position -= 20
                p.drawString(50, y_position, f"No Disponibles: {queryset.filter(disponible=False).count()}")
            
            p.showPage()
            p.save()
            
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            filename = f'reporte_{tipo}_con_estadisticas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            raise Exception(f"Error generando reporte PDF: {str(e)}")
